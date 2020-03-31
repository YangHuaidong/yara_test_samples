# -*- coding: utf-8 -*-

import os
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.writer.excel import ExcelWriter
# import pandas as pd
from common.config import Config
from common.md5 import md5_file
import xlrd

curdir = os.path.dirname(os.path.realpath(__file__))
cfg = Config()


def excel_result(data, md5_data, sample_type, sample_time):
    print("===========================")
    excel_log = os.path.join(curdir, "excel_log", sample_type)
    if not os.path.exists(excel_log):
        os.makedirs(excel_log)
    excel_name = sample_time + '.xls'
    excel_path = os.path.join(excel_log, excel_name)
    md5_list = []
    rulename_list = []
    black_excel = os.path.join(os.path.dirname(cfg.samplepath.black_path), 'black_sample.xls')
    if sample_type == 'white':
        for info in data:
            white_md5 = []
            white_md5.append(info['md5'])
            for rule in info["info"]:
                rulename = rule["meta"]["threatname"]
                if rulename not in white_md5:
                    white_md5.append(rulename)
                    rulename_list.append(white_md5)
        wb = Workbook()
        ws = wb.worksheets[0]
        ws.column_dimensions['A'].width = 40.0
        ws.column_dimensions['B'].width = 40.0
        ws.title = u"yara 误报工单"
        ws.append({1: u'MD5', 2: u'病毒名称'})
        for row in range(len(rulename_list)):
            ws.append(rulename_list[row])
            wb.save(filename=excel_path)
    else:
        file = xlrd.open_workbook(black_excel)
        sheet = file.sheets()[0]
        row = sheet.nrows
        col = sheet.ncols
        wb = Workbook()
        ws = wb.worksheets[0]
        ws.column_dimensions['A'].width = 40.0
        ws.column_dimensions['B'].width = 40.0
        ws.title = u"yara 未检出工单"
        ws.append({1: u'MD5', 2: u'病毒名称'})
        for info in data:
            md5_list.append(info['md5'])
        for md5 in md5_data:
            if md5 not in md5_list:
                for i in range(row):
                    if md5 in (sheet.row_values(i)):
                        ws.append(sheet.row_values(i))
        wb.save(filename=excel_path)
    return excel_path
        #         black_md5=[]
        #         demo_df=pd.read_excel(black_excel)  ##文件路径
        #         for indexs in demo_df.index:
        #             for i in range(len(demo_df.loc[indexs].values)):
        #                 if str(demo_df.loc[indexs].values[i]).rstrip() == md5:
        #                     rulename = demo_df.loc[indexs+1].values[i+1]
        #                     if rulename not in black_md5:
        #                         if md5 not in black_md5:
        #                             black_md5.append(md5)
        #                         black_md5.append(rulename)
        #         rulename_list.append(black_md5)
        # wb = Workbook()
        # ws = wb.worksheets[0]
        # ws.column_dimensions['A'].width = 40.0
        # ws.column_dimensions['B'].width = 40.0
        # ws.column_dimensions['C'].width = 40.0
        # ws.title = u"yara 未检出工单"
        # ws.append({1:u'MD5',2:u'病毒名称'})
        # for row in range(len(rulename_list)):
        #     ws.append(rulename_list[row])
        # wb.save(filename=excel_path)



if __name__ == '__main__':

    data2 = [{'info': [{'meta': {'comment': 'None', 'hacker': 'None', 'threattype': 'BackDoor', 'hash': '93d1a2e13a3368a2472043bd6331afe9', 'description': 'Semi-Auto-generated  - file Ajax_PHP Command Shell.php.txt', 'family': 'Ajax', 'author': 'Spider', 'date': 'None', 'judge': 'unknown', 'threatname': 'WebShell[BackDoor]/Unlimit.Ajax.PHP.Command.Shell.php.a', 'reference': 'None'}, 'rulename': 'WebShell_BackDoor_Unlimit_Ajax_PHP_Command_Shell_php_a'}, {'meta': {'comment': 'None', 'hacker': 'None', 'threattype': 'BackDoor', 'hash': '8bfa2eeb8a3ff6afc619258e39fded56', 'description': 'Semi-Auto-generated  - file ironshell.php.txt', 'family': 'ironshell', 'author': 'Spider', 'date': 'None', 'judge': 'unknown', 'threatname': 'WebShell[BackDoor]/Unlimit.ironshell.php.a', 'reference': 'None'}, 'rulename': 'WebShell_BackDoor_Unlimit_ironshell_php_a'}, {'meta': {'comment': 'None', 'hacker': 'None', 'threattype': 'BackDoor', 'description': 'PHP Webshells Github Archive - from files Ajax_PHP Command Shell.php, Ajax_PHP_Command_Shell.php, soldierofallah.php', 'family': 'WebShell', 'author': 'Spider', 'date': 'None', 'hash1': 'c0a4ba3e834fb63e0a220a43caaf55c654f97429', 'hash0': 'fa11deaee821ca3de7ad1caafa2a585ee1bc8d82', 'hash2': '16fa789b20409c1f2ffec74484a30d0491904064', 'judge': 'unknown', 'threatname': 'WebShell[BackDoor]/Unlimit.WebShell..Ajax.PHP.Command.Shell.Ajax.PHP.Command.Shell.soldierofallah.a', 'super_rule': 1, 'reference': 'None'}, 'rulename': 'WebShell_BackDoor_Unlimit_WebShell__Ajax_PHP_Command_Shell_Ajax_PHP_Command_Shell_soldierofallah_a'}], 'md5': '93d1a2e13a3368a2472043bd6331afe9'}, {'info': [{'meta': {'comment': 'None', 'hacker': 'None', 'threattype': 'BackDoor', 'hash': 'e6819b8f8ff2f1073f7d46a0b192f43b', 'description': 'Webshells Auto-generated - file admin-ad.asp', 'family': 'admin', 'author': 'Spider', 'date': 'None', 'judge': 'unknown', 'threatname': 'WebShell[BackDoor]/Unlimit.admin.ad.a', 'reference': 'None'}, 'rulename': 'WebShell_BackDoor_Unlimit_admin_ad_a'}, {'meta': {'comment': 'None', 'hacker': 'None', 'threattype': 'BackDoor', 'hash': '97f2552c2fafc0b2eb467ee29cc803c8', 'description': 'Webshells Auto-generated - file 2005.asp', 'family': 'FeliksPack3', 'author': 'Spider', 'date': 'None', 'judge': 'unknown', 'threatname': 'WebShell[BackDoor]/Unlimit.FeliksPack3...PHP.Shells.2005.a', 'reference': 'None'}, 'rulename': 'WebShell_BackDoor_Unlimit_FeliksPack3___PHP_Shells_2005_a'}], 'md5': 'e6819b8f8ff2f1073f7d46a0b192f43b'}, {'info': [{'meta': {'comment': 'None', 'hacker': 'None', 'threattype': 'BackDoor', 'hash': '7f83adcb4c1111653d30c6427a94f66f', 'description': 'Semi-Auto-generated  - file ak74shell.php.php.txt', 'family': 'ak74shell', 'author': 'Spider', 'date': 'None', 'judge': 'unknown', 'threatname': 'WebShell[BackDoor]/Unlimit.ak74shell.php.php.a', 'reference': 'None'}, 'rulename': 'WebShell_BackDoor_Unlimit_ak74shell_php_php_a'}, {'meta': {'comment': 'None', 'hacker': 'None', 'threattype': 'BackDoor', 'hash': 'c90b0ba575f432ecc08f8f292f3013b5532fe2c4', 'description': 'PHP Webshells Github Archive - file AK-74 Security Team Web Shell Beta Version.php', 'family': 'WebShell', 'author': 'Spider', 'date': 'None', 'judge': 'unknown', 'threatname': 'WebShell[BackDoor]/Unlimit.WebShell.AK.74.Security.Team.Web.Shell.Beta.Version.a', 'reference': 'None'}, 'rulename': 'WebShell_BackDoor_Unlimit_WebShell_AK_74_Security_Team_Web_Shell_Beta_Version_a'}], 'md5': '7f83adcb4c1111653d30c6427a94f66f'}, {'info': [{'meta': {'comment': 'None', 'hacker': 'None', 'threattype': 'BackDoor', 'hash': '40d0abceba125868be7f3f990f031521', 'description': 'Semi-Auto-generated  - file Antichat Shell v1.3.php.txt', 'family': 'Antichat', 'author': 'Spider', 'date': 'None', 'judge': 'unknown', 'threatname': 'WebShell[BackDoor]/Unlimit.Antichat.Shell.v1.3.php.a', 'reference': 'None'}, 'rulename': 'WebShell_BackDoor_Unlimit_Antichat_Shell_v1_3_php_a'}, {'meta': {'comment': 'None', 'hacker': 'None', 'threattype': 'BackDoor', 'hash': '9cfe372d49fe8bf2fac8e1c534153d9b', 'description': 'Semi-Auto-generated  - file Dx.php.php.txt', 'family': 'Dx', 'author': 'Spider', 'date': 'None', 'judge': 'unknown', 'threatname': 'WebShell[BackDoor]/Unlimit.Dx.php.php.a', 'reference': 'None'}, 'rulename': 'WebShell_BackDoor_Unlimit_Dx_php_php_a'}, {'meta': {'comment': 'None', 'hacker': 'None', 'threattype': 'BackDoor', 'hash': '12bbdf6ef403720442a47a3cc730d034', 'description': 'Semi-Auto-generated  - file mysql.php.php.txt', 'family': 'mysql', 'author': 'Spider', 'date': 'None', 'judge': 'unknown', 'threatname': 'WebShell[BackDoor]/Unlimit.mysql.php.php.a', 'reference': 'None'}, 'rulename': 'WebShell_BackDoor_Unlimit_mysql_php_php_a'}, {'meta': {'comment': 'None', 'hacker': 'None', 'threattype': 'BackDoor', 'hash': '5fbe4d8edeb2769eda5f4add9bab901e', 'description': 'Semi-Auto-generated  - file mysql_tool.php.php.txt', 'family': 'mysql', 'author': 'Spider', 'date': 'None', 'judge': 'unknown', 'threatname': 'WebShell[BackDoor]/Unlimit.mysql.tool.php.php.a', 'reference': 'None'}, 'rulename': 'WebShell_BackDoor_Unlimit_mysql_tool_php_php_a'}, {'meta': {'comment': 'None', 'hacker': 'None', 'threattype': 'BackDoor', 'hash': '40d0abceba125868be7f3f990f031521', 'description': 'Web Shell - file Antichat Shell v1.3.php', 'family': 'webshell', 'author': 'Spider', 'date': '2014/01/28', 'score': 70, 'judge': 'unknown', 'threatname': 'WebShell[BackDoor]/Unlimit.webshell.Antichat.Shell.v1.3.2.a', 'reference': 'None'}, 'rulename': 'WebShell_BackDoor_Unlimit_webshell_Antichat_Shell_v1_3_2_a'}, {'meta': {'comment': 'None', 'hacker': 'None', 'threattype': 'BackDoor', 'description': 'PHP Webshells Github Archive - from files Antichat Shell v1.3.php, Antichat Shell. Modified by Go0o$E.php, Antichat Shell.php, fatal.php', 'family': 'WebShell', 'author': 'Spider', 'date': 'None', 'hash1': 'd710c95d9f18ec7c76d9349a28dd59c3605c02be', 'hash0': 'd829e87b3ce34460088c7775a60bded64e530cd4', 'hash3': '41780a3e8c0dc3cbcaa7b4d3c066ae09fb74a289', 'hash2': 'f044d44e559af22a1a7f9db72de1206f392b8976', 'judge': 'unknown', 'threatname': 'WebShell[BackDoor]/Unlimit.WebShell.Generic.PHP.3.a', 'super_rule': 1, 'reference': 'None'}, 'rulename': 'WebShell_BackDoor_Unlimit_WebShell_Generic_PHP_3_a'}], 'md5': '40d0abceba125868be7f3f990f031521'}, {'info': [{'meta': {'comment': 'None', 'hacker': 'None', 'threattype': 'BackDoor', 'hash': 'b6f468252407efc2318639da22b08af0', 'description': 'Semi-Auto-generated  - file Ajan.asp.txt', 'family': 'Ajan', 'author': 'Spider', 'date': 'None', 'judge': 'unknown', 'threatname': 'WebShell[BackDoor]/Unlimit.Ajan.asp.a', 'reference': 'None'}, 'rulename': 'WebShell_BackDoor_Unlimit_Ajan_asp_a'}, {'meta': {'comment': 'None', 'hacker': 'None', 'threattype': 'BackDoor', 'hash': '22194f8c44524f80254e1b5aec67b03e', 'description': 'Webshells Auto-generated - file ajan.asp', 'family': 'FSO', 'author': 'Spider', 'date': 'None', 'judge': 'unknown', 'threatname': 'WebShell[BackDoor]/Unlimit.FSO.s.ajan.2.a', 'reference': 'None'}, 'rulename': 'WebShell_BackDoor_Unlimit_FSO_s_ajan_2_a'}, {'meta': {'comment': 'None', 'hacker': 'None', 'threattype': 'BackDoor', 'hash': '22194f8c44524f80254e1b5aec67b03e', 'description': 'Webshells Auto-generated - file ajan.asp', 'family': 'FSO', 'author': 'Spider', 'date': 'None', 'judge': 'unknown', 'threatname': 'WebShell[BackDoor]/Unlimit.FSO.s.ajan.a', 'reference': 'None'}, 'rulename': 'WebShell_BackDoor_Unlimit_FSO_s_ajan_a'}, {'meta': {'comment': 'None', 'hacker': 'None', 'threattype': 'BackDoor', 'hash': 'b6f468252407efc2318639da22b08af0', 'description': 'Web Shell - file Ajan.asp', 'family': 'webshell', 'author': 'Spider', 'date': '2014/01/28', 'score': 70, 'judge': 'unknown', 'threatname': 'WebShell[BackDoor]/Unlimit.webshell.asp.Ajan.a', 'reference': 'None'}, 'rulename': 'WebShell_BackDoor_Unlimit_webshell_asp_Ajan_a'}], 'md5': 'b6f468252407efc2318639da22b08af0'}]
    md51_data = ['0684a64086ad1114949a1e51f06aa750', '7e87156e96818a72d5c41e5e3fda654f', '6a57f7c43c2bad2682dd83448cb52d81', 'dc6900546369047d083fb0dace072d14', '8995861be722bf51841456f671c9b691', '77ad75213462f202d014f50e2c127ad6', '578f64b61cf47b7083e8d4c95638007c', 'ce9d36e3ebff58fb8e9b27c75e172517', '0dc254d9ad4b70a3f2e4030e02498f4e', '5e2863454df5de4b3e723caf330c3438', '321ca1832a700ba04caa6d0f3aa15ccf', '93d1a2e13a3368a2472043bd6331afe9', 'e6819b8f8ff2f1073f7d46a0b192f43b', '05661dc1c90ad70ba5eb4efa3c66b457', '7f83adcb4c1111653d30c6427a94f66f', '4664aa9c6eb0259f818753f8470c7a83', 'c05421f3295b96ec7af7c2fc6c375a00', 'a7aa60644345470eb94e0b936eddb1ba', 'c253c48252c640de1cf24e9433a40d78', '943780214cb08970b8271eb9f8f874f5', 'd9c057125c8391edebc0c37372b105ba', 'eab3cb53d811ec74be1f1c6323d5d23f', '6bca6bc809169f86f7f2ece19fcfc293', '6d37d81ce9431c4cfcc828c535acb5c7', '06f7d36934bfc9ecc200b2b093752c08', 'f3a3156f3816005289d0377de7009f05', '048c36c7f0f691f437aa03befc34762b', '5e9c9aaa7dabc9e92abd02ce7df59fdf', '6e62c974334fd28376c73fc84f318b34', 'd936dec4e0431156fbec5baa31f86f9d', 'c92681a97165f3fb9eb441ddffac6237', '40d0abceba125868be7f3f990f031521', '8995b89c7ec2cd51ac974cbad169ed5c', 'b6f468252407efc2318639da22b08af0', '3fc36f387276937f5d15b3a0cc3791f8', 'acf1a8d2cc4e7f3303f947bec4f78130', 'a94a00ef4f90597ecd883b1ca3a68a1d']
    excel_result(data2, md51_data, 'black', '20190011-100000')
